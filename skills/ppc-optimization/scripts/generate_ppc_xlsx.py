#!/usr/bin/env python3
"""
XLSX Report Generator for PPC Optimization
Creates a multi-sheet Excel report with complete PPC analysis and recommendations.
"""

import json
import argparse
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl', '--break-system-packages', '-q'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_headers(ws, row, count):
    for col in range(1, count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def classification_fill(classification):
    fills = {
        'winner': GREEN_FILL,
        'marginal': YELLOW_FILL,
        'bleeder': RED_FILL,
        'wasted_spend': RED_FILL,
        'under_review': YELLOW_FILL,
    }
    return fills.get(classification, PatternFill())


def create_dashboard_sheet(wb, search_terms_data, campaign_structure):
    """Sheet 1: PPC Dashboard — KPI overview."""
    ws = wb.active
    ws.title = "PPC Dashboard"

    # KPI Summary
    ws.cell(row=1, column=1, value='PPC PERFORMANCE DASHBOARD').font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    kpis = [
        ('Total Spend', f"${search_terms_data.get('total_spend', 0):,.2f}"),
        ('Total Sales', f"${search_terms_data.get('total_sales', 0):,.2f}"),
        ('Overall ACOS', f"{search_terms_data.get('overall_acos', 0):.1f}%"),
        ('Unique Search Terms', str(search_terms_data.get('total_unique_terms', 0))),
    ]

    for i, (label, value) in enumerate(kpis):
        ws.cell(row=3, column=i*2+1, value=label).font = Font(bold=True, size=11)
        cell = ws.cell(row=4, column=i*2+1, value=value)
        cell.font = Font(bold=True, size=14)

    # Classification summary
    ws.cell(row=6, column=1, value='SEARCH TERM CLASSIFICATION').font = Font(bold=True, size=13)
    class_summary = search_terms_data.get('classification_summary', {})

    headers = ['Classification', 'Count', 'Action']
    for col, h in enumerate(headers, 1):
        ws.cell(row=7, column=col, value=h)
    style_headers(ws, 7, len(headers))

    class_info = [
        ('Winners', class_summary.get('winner', 0), 'Scale up — increase bids/budget'),
        ('Marginal', class_summary.get('marginal', 0), 'Optimize bids to hit target ACOS'),
        ('Bleeders', class_summary.get('bleeder', 0), 'Reduce bids or pause'),
        ('Wasted Spend', class_summary.get('wasted_spend', 0), 'Add as negative keywords'),
        ('Under Review', class_summary.get('under_review', 0), 'Monitor — need more data'),
        ('Insufficient Data', class_summary.get('insufficient_data', 0), 'Increase bids for more impressions'),
    ]

    for i, (name, count, action) in enumerate(class_info, 8):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=count)
        ws.cell(row=i, column=3, value=action)

    # Recommended budget
    if campaign_structure:
        row = 15
        ws.cell(row=row, column=1, value='RECOMMENDED BUDGET').font = Font(bold=True, size=13)
        ws.cell(row=row+1, column=1, value='Daily Budget:').font = Font(bold=True)
        ws.cell(row=row+1, column=2, value=f"${campaign_structure.get('total_daily_budget', 0):.2f}")
        ws.cell(row=row+2, column=1, value='Monthly Budget:').font = Font(bold=True)
        ws.cell(row=row+2, column=2, value=f"${campaign_structure.get('total_monthly_budget', 0):.2f}")
        ws.cell(row=row+3, column=1, value='Target ACOS:').font = Font(bold=True)
        ws.cell(row=row+3, column=2, value=f"{campaign_structure.get('target_acos', 30)}%")

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15


def create_keyword_strategy_sheet(wb, keyword_priorities):
    """Sheet 2: Keyword Strategy — all keywords with tiers and bids."""
    ws = wb.create_sheet("Keyword Strategy")

    headers = ['Keyword', 'Search Volume', 'Tier', 'Match Type', 'PPC Orders',
               'PPC ACOS', 'PPC CVR', 'PPC Spend', 'Priority Score', 'Source']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_headers(ws, 1, len(headers))

    tier_fills = {
        'tier_1_exact': GREEN_FILL,
        'tier_2_phrase': BLUE_FILL,
        'tier_3_broad': YELLOW_FILL,
        'negative_candidate': RED_FILL,
    }

    keywords = keyword_priorities.get('keywords', [])
    for i, kw in enumerate(keywords[:300], 2):
        ws.cell(row=i, column=1, value=kw.get('keyword', ''))
        ws.cell(row=i, column=2, value=kw.get('search_volume', 0))
        tier = kw.get('tier', '')
        cell = ws.cell(row=i, column=3, value=tier.replace('_', ' ').title())
        cell.fill = tier_fills.get(tier, PatternFill())
        ws.cell(row=i, column=4, value=kw.get('recommended_match_type', ''))
        ws.cell(row=i, column=5, value=kw.get('ppc_orders', 0))
        ws.cell(row=i, column=6, value=f"{kw.get('ppc_acos', 0):.1f}%")
        ws.cell(row=i, column=7, value=f"{kw.get('ppc_cvr', 0):.1f}%")
        ws.cell(row=i, column=8, value=f"${kw.get('ppc_spend', 0):.2f}")
        ws.cell(row=i, column=9, value=kw.get('ppc_priority_score', 0))
        ws.cell(row=i, column=10, value=kw.get('source', ''))

    widths = [35, 14, 18, 14, 12, 12, 12, 12, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_search_term_sheet(wb, search_terms_data):
    """Sheet 3: Search Term Analysis — full report with classifications."""
    ws = wb.create_sheet("Search Term Analysis")

    headers = ['Search Term', 'Impressions', 'Clicks', 'CTR', 'CPC',
               'Spend', 'Sales', 'Orders', 'ACOS', 'CVR', 'Classification', 'Action']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_headers(ws, 1, len(headers))

    terms = search_terms_data.get('search_terms', [])
    for i, term in enumerate(terms[:500], 2):
        ws.cell(row=i, column=1, value=term.get('search_term', ''))
        ws.cell(row=i, column=2, value=term.get('impressions', 0))
        ws.cell(row=i, column=3, value=term.get('clicks', 0))
        ctr = (term.get('clicks', 0) / term.get('impressions', 1) * 100) if term.get('impressions', 0) > 0 else 0
        ws.cell(row=i, column=4, value=f"{ctr:.2f}%")
        ws.cell(row=i, column=5, value=f"${term.get('cpc', 0):.2f}")
        ws.cell(row=i, column=6, value=f"${term.get('spend', 0):.2f}")
        ws.cell(row=i, column=7, value=f"${term.get('sales', 0):.2f}")
        ws.cell(row=i, column=8, value=term.get('orders', 0))
        ws.cell(row=i, column=9, value=f"{term.get('acos', 0):.1f}%")
        ws.cell(row=i, column=10, value=f"{term.get('cvr', 0):.1f}%")

        classification = term.get('classification', '')
        cell = ws.cell(row=i, column=11, value=classification.replace('_', ' ').title())
        cell.fill = classification_fill(classification)

        ws.cell(row=i, column=12, value=term.get('action', '').replace('_', ' ').title())

    widths = [35, 12, 10, 8, 8, 10, 10, 8, 8, 8, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_campaign_structure_sheet(wb, campaign_structure):
    """Sheet 4: Campaign Structure — recommended campaigns and ad groups."""
    ws = wb.create_sheet("Campaign Structure")

    if not campaign_structure:
        ws.cell(row=1, column=1, value='No campaign structure data available')
        return

    ws.cell(row=1, column=1, value='RECOMMENDED CAMPAIGN STRUCTURE').font = Font(bold=True, size=14)

    row = 3
    for campaign in campaign_structure.get('campaigns', []):
        # Campaign header
        ws.cell(row=row, column=1, value=campaign.get('campaign_name', '')).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value=f"Type: {campaign.get('campaign_type', '')}")
        ws.cell(row=row, column=3, value=f"Budget: ${campaign.get('daily_budget', 0):.2f}/day")
        ws.cell(row=row, column=4, value=f"Strategy: {campaign.get('bidding_strategy', '')}")
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = BLUE_FILL
        row += 1

        # Ad group headers
        headers = ['Ad Group', 'Keyword', 'Match Type', 'Bid', 'Search Volume', 'Notes']
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_headers(ws, row, len(headers))
        row += 1

        for ag in campaign.get('ad_groups', []):
            ag_name = ag.get('ad_group_name', '')
            keywords = ag.get('keywords', [])

            if keywords:
                for kw in keywords:
                    ws.cell(row=row, column=1, value=ag_name)
                    ws.cell(row=row, column=2, value=kw.get('keyword', ''))
                    ws.cell(row=row, column=3, value=kw.get('match_type', ''))
                    ws.cell(row=row, column=4, value=f"${kw.get('bid', 0):.2f}")
                    ws.cell(row=row, column=5, value=kw.get('search_volume', 0))
                    row += 1
            else:
                ws.cell(row=row, column=1, value=ag_name)
                ws.cell(row=row, column=2, value=ag.get('targeting_type', 'auto'))
                ws.cell(row=row, column=4, value=f"${ag.get('default_bid', 0):.2f}")
                ws.cell(row=row, column=6, value=ag.get('note', ''))
                row += 1

        row += 1  # Space between campaigns

    # Cross-campaign negatives
    row += 1
    ws.cell(row=row, column=1, value='CROSS-CAMPAIGN NEGATIVES').font = Font(bold=True, size=12)
    row += 1
    cross_negs = campaign_structure.get('cross_campaign_negatives', {})
    ws.cell(row=row, column=1, value=cross_negs.get('description', ''))
    row += 1
    for kw in cross_negs.get('keywords', []):
        ws.cell(row=row, column=1, value=kw)
        ws.cell(row=row, column=2, value='negative exact')
        row += 1

    widths = [25, 35, 14, 10, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_negatives_sheet(wb, negatives_data, campaign_structure):
    """Sheet 5: Negative Keywords."""
    ws = wb.create_sheet("Negative Keywords")

    headers = ['Keyword', 'Match Type', 'Reason', 'Spend', 'Type']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_headers(ws, 1, len(headers))

    row = 2

    # From campaign structure negative candidates
    if campaign_structure:
        for neg in campaign_structure.get('negative_keyword_candidates', []):
            ws.cell(row=row, column=1, value=neg.get('keyword', ''))
            ws.cell(row=row, column=2, value=neg.get('match_type', 'negative_exact'))
            ws.cell(row=row, column=3, value=neg.get('reason', ''))
            ws.cell(row=row, column=4, value='').fill = RED_FILL
            ws.cell(row=row, column=5, value='Wasted Spend')
            row += 1

        # Cross-campaign negatives
        for kw in campaign_structure.get('cross_campaign_negatives', {}).get('keywords', []):
            ws.cell(row=row, column=1, value=kw)
            ws.cell(row=row, column=2, value='negative_exact')
            ws.cell(row=row, column=3, value='Prevent cannibalization in Research/Auto campaigns')
            ws.cell(row=row, column=5, value='Cross-Campaign')
            row += 1

    widths = [35, 18, 50, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_action_items_sheet(wb, search_terms_data, keyword_priorities, campaign_structure):
    """Sheet 6: Prioritized Action Items."""
    ws = wb.create_sheet("Action Items")

    headers = ['Priority', 'Action', 'Details', 'Expected Impact']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_headers(ws, 1, len(headers))

    actions = []

    # Check for wasted spend
    wasted = search_terms_data.get('classification_summary', {}).get('wasted_spend', 0)
    if wasted > 0:
        wasted_terms = [t for t in search_terms_data.get('search_terms', [])
                        if t.get('classification') == 'wasted_spend']
        wasted_spend = sum(t.get('spend', 0) for t in wasted_terms)
        actions.append(('High', 'Add negative keywords',
                        f'{wasted} terms with ${wasted_spend:.2f} wasted spend',
                        f'Save ~${wasted_spend:.2f}/month'))

    # Check for winners to scale
    winners = search_terms_data.get('classification_summary', {}).get('winner', 0)
    if winners > 0:
        actions.append(('High', 'Scale winning keywords',
                        f'{winners} profitable keywords — increase bids 10-20%',
                        'Higher impression share on proven converters'))

    # Check for bleeders
    bleeders = search_terms_data.get('classification_summary', {}).get('bleeder', 0)
    if bleeders > 0:
        actions.append(('Medium', 'Reduce bids on bleeders',
                        f'{bleeders} keywords above target ACOS — reduce bids 20-30%',
                        'Lower ACOS while maintaining some volume'))

    # Campaign structure recommendation
    if campaign_structure:
        actions.append(('Medium', 'Implement campaign structure',
                        f"{campaign_structure.get('campaign_count', 0)} campaigns, "
                        f"${campaign_structure.get('total_daily_budget', 0):.2f}/day budget",
                        'Better keyword control and budget allocation'))

    # Cross-campaign negatives
    if campaign_structure and campaign_structure.get('cross_campaign_negatives', {}).get('keywords'):
        count = len(campaign_structure['cross_campaign_negatives']['keywords'])
        actions.append(('Medium', 'Add cross-campaign negatives',
                        f'{count} exact match winners to negate in broad/phrase campaigns',
                        'Prevent keyword cannibalization'))

    priority_fills = {'High': RED_FILL, 'Medium': YELLOW_FILL, 'Low': GREEN_FILL}
    for i, (priority, action, details, impact) in enumerate(actions, 2):
        ws.cell(row=i, column=1, value=priority).fill = priority_fills.get(priority, YELLOW_FILL)
        ws.cell(row=i, column=2, value=action)
        ws.cell(row=i, column=3, value=details).alignment = WRAP
        ws.cell(row=i, column=4, value=impact).alignment = WRAP

    widths = [12, 30, 50, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_xlsx(search_terms_path, keyword_priorities_path,
                  campaign_structure_path, output_path,
                  performance_path=None, negatives_path=None, bid_strategy_path=None):
    """Main XLSX generation."""

    with open(search_terms_path) as f:
        search_terms = json.load(f)

    keyword_priorities = {}
    if keyword_priorities_path and os.path.exists(keyword_priorities_path):
        with open(keyword_priorities_path) as f:
            keyword_priorities = json.load(f)

    campaign_structure = {}
    if campaign_structure_path and os.path.exists(campaign_structure_path):
        with open(campaign_structure_path) as f:
            campaign_structure = json.load(f)

    wb = Workbook()

    create_dashboard_sheet(wb, search_terms, campaign_structure)
    create_keyword_strategy_sheet(wb, keyword_priorities)
    create_search_term_sheet(wb, search_terms)
    create_campaign_structure_sheet(wb, campaign_structure)
    create_negatives_sheet(wb, {}, campaign_structure)
    create_action_items_sheet(wb, search_terms, keyword_priorities, campaign_structure)

    wb.save(output_path)
    print(f"\nPPC Optimization XLSX saved to: {output_path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate PPC Optimization XLSX report')
    parser.add_argument('--search-terms', required=True)
    parser.add_argument('--keyword-priorities', required=True)
    parser.add_argument('--performance', default=None)
    parser.add_argument('--campaign-structure', required=True)
    parser.add_argument('--bid-strategy', default=None)
    parser.add_argument('--negatives', default=None)
    parser.add_argument('--output', required=True)
    args = parser.parse_args()

    generate_xlsx(args.search_terms, args.keyword_priorities,
                  args.campaign_structure, args.output,
                  args.performance, args.negatives, args.bid_strategy)
