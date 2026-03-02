"""
Generate Competitor Research XLSX for Zoviro.
Creates a professional Excel workbook with 4 sheets:
  1. Niche Overview - market summary per keyword
  2. All Competitors - full deduplicated list
  3. Top Competitors - detailed top 15-20
  4. Feature Comparison - Zoviro vs Top 5 side-by-side
"""
import json
import os
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

WORKSPACE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

# Style constants
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=10)
DATA_FONT = Font(name='Arial', size=10)
MONEY_FORMAT = '$#,##0.00'
MONEY_FORMAT_K = '$#,##0'
INT_FORMAT = '#,##0'
PCT_FORMAT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
ZOVIRO_FILL = PatternFill('solid', fgColor='E2EFDA')
TOP5_FILL = PatternFill('solid', fgColor='FCE4D6')


def style_header_row(ws, row_num, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, value=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical='center')
    if fmt:
        cell.number_format = fmt
    return cell


def auto_width(ws, min_width=8, max_width=40):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                max_len = max(max_len, min(cell_len + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def create_niche_overview(wb, niche_data):
    """Sheet 1: Niche Overview."""
    ws = wb.active
    ws.title = "Niche Overview"

    headers = ['Search Keyword', 'Search Volume', 'Page 1 Products', 'Average BSR',
               'Average Price', 'Average Rating', 'Average Reviews', 'Price Range']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    for i, niche in enumerate(niche_data, 2):
        style_data_cell(ws, i, 1, niche.get('keyword'))
        style_data_cell(ws, i, 2, niche.get('search_volume'), INT_FORMAT)
        style_data_cell(ws, i, 3, niche.get('page1_products'), INT_FORMAT)
        style_data_cell(ws, i, 4, niche.get('avg_bsr'), INT_FORMAT)
        style_data_cell(ws, i, 5, niche.get('avg_price'), MONEY_FORMAT)
        style_data_cell(ws, i, 6, niche.get('avg_rating'), '0.0')
        style_data_cell(ws, i, 7, niche.get('avg_reviews'), INT_FORMAT)
        style_data_cell(ws, i, 8, niche.get('price_range'))

    auto_width(ws)
    ws.freeze_panes = 'A2'
    return ws


def create_all_competitors(wb, merged_data):
    """Sheet 2: All Competitors (full list)."""
    ws = wb.create_sheet("All Competitors")

    headers = ['#', 'ASIN', 'Title', 'Brand', 'Price', 'BSR', 'Rating',
               'Reviews', 'Keywords Found In', 'Keyword Freq', 'Relevance Score', 'URL']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    competitors = merged_data.get('competitors', [])
    for i, c in enumerate(competitors, 2):
        style_data_cell(ws, i, 1, i - 1)
        style_data_cell(ws, i, 2, c.get('asin'))
        style_data_cell(ws, i, 3, (c.get('title', '')[:80] + '...') if len(c.get('title', '')) > 80 else c.get('title'))
        style_data_cell(ws, i, 4, c.get('brand'))
        style_data_cell(ws, i, 5, c.get('price'), MONEY_FORMAT)
        style_data_cell(ws, i, 6, c.get('bsr'), INT_FORMAT)
        style_data_cell(ws, i, 7, c.get('rating'), '0.0')
        style_data_cell(ws, i, 8, c.get('review_count'), INT_FORMAT)
        style_data_cell(ws, i, 9, ', '.join(c.get('keyword_sources', [])))
        style_data_cell(ws, i, 10, c.get('keyword_frequency'))
        style_data_cell(ws, i, 11, c.get('relevance_score'), '0.0')
        style_data_cell(ws, i, 12, c.get('url'))

    auto_width(ws, max_width=50)
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['L'].width = 40
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:L{len(competitors) + 1}"
    return ws


def create_top_competitors(wb, top_data):
    """Sheet 3: Top Competitors Detail."""
    ws = wb.create_sheet("Top Competitors")

    headers = ['Rank', 'ASIN', 'Title', 'Brand', 'Price', 'Price/Unit',
               'BSR (Overall)', 'BSR (Subcategory)', 'Rating', 'Reviews',
               '30-Day Revenue', '30-Day Units', 'Listing Health',
               'Sellers', 'Fulfillment', 'Images', 'A+ Content', 'Video',
               'Keyword Freq', 'Relevance Score', 'Top Keywords']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    competitors = top_data.get('competitors', [])
    for i, c in enumerate(competitors, 2):
        style_data_cell(ws, i, 1, i - 1)
        style_data_cell(ws, i, 2, c.get('asin'))
        title = c.get('title', '')
        style_data_cell(ws, i, 3, (title[:70] + '...') if len(title) > 70 else title)
        style_data_cell(ws, i, 4, c.get('brand'))
        style_data_cell(ws, i, 5, c.get('price'), MONEY_FORMAT)
        style_data_cell(ws, i, 6, c.get('price_per_unit', ''))
        style_data_cell(ws, i, 7, c.get('bsr'), INT_FORMAT)
        bsr_sub = c.get('bsr_subcategory')
        style_data_cell(ws, i, 8, bsr_sub if bsr_sub else '', INT_FORMAT if bsr_sub else None)
        style_data_cell(ws, i, 9, c.get('rating'), '0.0')
        style_data_cell(ws, i, 10, c.get('review_count'), INT_FORMAT)
        rev = c.get('revenue_30d')
        style_data_cell(ws, i, 11, rev if rev else '', MONEY_FORMAT_K if rev else None)
        units = c.get('units_30d')
        style_data_cell(ws, i, 12, units if units else '', INT_FORMAT if units else None)
        lhs = c.get('listing_health_score')
        style_data_cell(ws, i, 13, lhs if lhs else '')
        style_data_cell(ws, i, 14, c.get('seller_count', ''))
        style_data_cell(ws, i, 15, c.get('fulfillment', ''))
        style_data_cell(ws, i, 16, c.get('image_count', ''))
        style_data_cell(ws, i, 17, 'Yes' if c.get('has_aplus') else ('No' if c.get('detail_collected') else ''))
        style_data_cell(ws, i, 18, 'Yes' if c.get('has_video') else ('No' if c.get('detail_collected') else ''))
        style_data_cell(ws, i, 19, c.get('keyword_frequency'))
        style_data_cell(ws, i, 20, c.get('relevance_score'), '0.0')
        kws = c.get('top_keywords', [])
        style_data_cell(ws, i, 21, ', '.join(kws[:5]) if kws else '')

    auto_width(ws, max_width=45)
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['U'].width = 40
    ws.freeze_panes = 'A2'
    return ws


def create_feature_comparison(wb, our_product, top_competitors):
    """Sheet 4: Feature Comparison - Zoviro vs Top 5."""
    ws = wb.create_sheet("Feature Comparison")

    top5 = top_competitors[:5]
    num_cols = 2 + len(top5)

    # Header row with product names
    ws.cell(row=1, column=1, value='Feature')
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL

    ws.cell(row=1, column=2, value=f"Zoviro ({our_product.get('asin', '')})")
    ws.cell(row=1, column=2).font = HEADER_FONT
    ws.cell(row=1, column=2).fill = PatternFill('solid', fgColor='548235')

    for i, c in enumerate(top5):
        col = i + 3
        ws.cell(row=1, column=col, value=f"{c.get('brand', 'N/A')} ({c.get('asin', '')})")
        ws.cell(row=1, column=col).font = HEADER_FONT
        ws.cell(row=1, column=col).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # Feature rows
    features = [
        ('Product Title', 'title', None),
        ('Brand', 'brand', None),
        ('Price', 'price', MONEY_FORMAT),
        ('Price Per Unit', 'price_per_unit', None),
        ('Rating', 'rating', '0.0'),
        ('Review Count', 'review_count', INT_FORMAT),
        ('BSR (Overall)', 'bsr', INT_FORMAT),
        ('BSR (Subcategory)', 'bsr_subcategory', INT_FORMAT),
        ('30-Day Revenue', 'revenue_30d', MONEY_FORMAT_K),
        ('30-Day Units', 'units_30d', INT_FORMAT),
        ('Listing Health Score', 'listing_health_score', None),
        ('Fulfillment', 'fulfillment', None),
        ('Seller Count', 'seller_count', None),
        ('Image Count', 'image_count', None),
        ('A+ Content', 'has_aplus', None),
        ('Video', 'has_video', None),
        ('Keyword Overlap', 'keyword_frequency', None),
    ]

    for row_idx, (label, key, fmt) in enumerate(features, 2):
        cell = ws.cell(row=row_idx, column=1, value=label)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

        # Zoviro column
        val = our_product.get(key, '')
        if key == 'has_aplus':
            val = 'Yes' if val else 'No' if val is not None else ''
        if key == 'has_video':
            val = 'Yes' if val else 'No' if val is not None else ''
        c = style_data_cell(ws, row_idx, 2, val, fmt)
        c.fill = ZOVIRO_FILL

        # Competitor columns
        for i, comp in enumerate(top5):
            val = comp.get(key, '')
            if key == 'has_aplus':
                val = 'Yes' if val else 'No' if val is not None else ''
            if key == 'has_video':
                val = 'Yes' if val else 'No' if val is not None else ''
            style_data_cell(ws, row_idx, i + 3, val, fmt)

    auto_width(ws, min_width=15, max_width=40)
    ws.column_dimensions['A'].width = 20
    ws.freeze_panes = 'B2'
    return ws


def generate_competitor_xlsx(merged_path, top_path, our_asin, config_path, output_path, niche_data=None):
    """Main function to generate the competitor research XLSX."""
    with open(merged_path) as f:
        merged = json.load(f)
    with open(top_path) as f:
        top = json.load(f)
    with open(config_path) as f:
        config = json.load(f)

    # Find our product info
    our_product = {}
    products = config.get('products', {})
    if our_asin in products:
        our_product = {
            'asin': our_asin,
            'title': products[our_asin].get('name', ''),
            'brand': config.get('brand', 'Zoviro'),
        }

    # Also check if our ASIN appears in any competitor data (for BSR/price/etc)
    for c in merged.get('competitors', []):
        if c.get('asin') == our_asin:
            our_product.update(c)
            break

    if niche_data is None:
        niche_data = []

    wb = Workbook()

    # Sheet 1: Niche Overview
    if niche_data:
        create_niche_overview(wb, niche_data)
    else:
        ws = wb.active
        ws.title = "Niche Overview"
        ws.cell(row=1, column=1, value="No niche data collected")
        ws.cell(row=2, column=1, value="Run the full skill workflow to populate this sheet")

    # Sheet 2: All Competitors
    create_all_competitors(wb, merged)

    # Sheet 3: Top Competitors
    create_top_competitors(wb, top)

    # Sheet 4: Feature Comparison
    top_competitors = top.get('competitors', [])
    create_feature_comparison(wb, our_product, top_competitors)

    # Add metadata sheet (hidden)
    ws_meta = wb.create_sheet("_metadata")
    ws_meta.cell(row=1, column=1, value="Generated")
    ws_meta.cell(row=1, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
    ws_meta.cell(row=2, column=1, value="Our ASIN")
    ws_meta.cell(row=2, column=2, value=our_asin)
    ws_meta.cell(row=3, column=1, value="Total Competitors")
    ws_meta.cell(row=3, column=2, value=merged.get('unique_competitors', 0))
    ws_meta.cell(row=4, column=1, value="Top N Analyzed")
    ws_meta.cell(row=4, column=2, value=top.get('top_n_selected', 0))
    ws_meta.sheet_state = 'hidden'

    wb.save(output_path)
    print(f"XLSX saved to: {output_path}")
    print(f"  Sheet 1: Niche Overview ({len(niche_data)} keywords)")
    print(f"  Sheet 2: All Competitors ({merged.get('unique_competitors', 0)} rows)")
    print(f"  Sheet 3: Top Competitors ({top.get('top_n_selected', 0)} rows)")
    print(f"  Sheet 4: Feature Comparison (Zoviro vs Top 5)")
    return output_path


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate competitor research XLSX')
    parser.add_argument('--merged', required=True)
    parser.add_argument('--detailed', required=True)
    parser.add_argument('--our-asin', required=True)
    parser.add_argument('--config', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--niche-data', help='Path to niche data JSON')
    args = parser.parse_args()

    merged_path = os.path.join(WORKSPACE, args.merged) if not os.path.isabs(args.merged) else args.merged
    detailed_path = os.path.join(WORKSPACE, args.detailed) if not os.path.isabs(args.detailed) else args.detailed
    config_path = os.path.join(WORKSPACE, args.config) if not os.path.isabs(args.config) else args.config
    output_path = os.path.join(WORKSPACE, args.output) if not os.path.isabs(args.output) else args.output

    niche_data = None
    if args.niche_data:
        niche_path = os.path.join(WORKSPACE, args.niche_data) if not os.path.isabs(args.niche_data) else args.niche_data
        with open(niche_path) as f:
            niche_data = json.load(f)

    generate_competitor_xlsx(merged_path, detailed_path, args.our_asin, config_path, output_path, niche_data)
