#!/usr/bin/env python3
"""
XLSX Report Generator for Listing Optimization
Creates a multi-sheet Excel report with analysis, comparison, and optimized copy.
"""

import json
import argparse
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl', '--break-system-packages', '-q'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SCORE_HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
SCORE_MED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SCORE_LOW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header_row(ws, row, col_count):
    """Apply header styling to a row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def score_fill(score):
    """Return fill color based on score."""
    if score >= 70:
        return SCORE_HIGH_FILL
    elif score >= 40:
        return SCORE_MED_FILL
    else:
        return SCORE_LOW_FILL


def create_scorecard_sheet(wb, gap_analysis):
    """Sheet 1: Listing Scorecard — our score vs benchmarks."""
    ws = wb.active
    ws.title = "Listing Scorecard"

    headers = ['Dimension', 'Our Score', 'Weight', 'Weighted Score', 'Status', 'Key Finding']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    scores = gap_analysis.get('scores', {})
    dimensions = [
        ('Title', scores.get('title', {}).get('score', 0), '30%',
         f"Missing keywords: {', '.join(scores.get('title', {}).get('keywords_missing', [])[:3])}"),
        ('Bullet Points', scores.get('bullets', {}).get('score', 0), '25%',
         f"{scores.get('bullets', {}).get('our_bullet_count', 0)} bullets, "
         f"avg {scores.get('bullets', {}).get('our_avg_bullet_length', 0)} chars"),
        ('Images', scores.get('images', {}).get('score', 0), '20%',
         f"{scores.get('images', {}).get('our_image_count', 0)} images "
         f"(avg competitor: {scores.get('images', {}).get('avg_competitor_image_count', 0)})"),
        ('A+ Content', scores.get('aplus', {}).get('score', 0), '15%',
         f"{'Has A+' if scores.get('aplus', {}).get('our_has_aplus') else 'No A+'} — "
         f"{int(scores.get('aplus', {}).get('competitor_aplus_ratio', 0)*100)}% competitors have it"),
        ('Backend Keywords', min(100, len(gap_analysis.get('backend_keywords', {}).get('selected_for_backend', [])) * 5), '10%',
         f"{gap_analysis.get('backend_keywords', {}).get('total_bytes_used', 0)}/250 bytes used"),
    ]

    for i, (dim, score, weight, finding) in enumerate(dimensions, 2):
        ws.cell(row=i, column=1, value=dim)
        cell = ws.cell(row=i, column=2, value=score)
        cell.fill = score_fill(score)
        ws.cell(row=i, column=3, value=weight)
        weight_num = float(weight.replace('%', '')) / 100
        ws.cell(row=i, column=4, value=round(score * weight_num, 1))
        ws.cell(row=i, column=5, value='Good' if score >= 70 else 'Needs Work' if score >= 40 else 'Critical')
        ws.cell(row=i, column=6, value=finding)

    # Overall score row
    row = len(dimensions) + 2
    ws.cell(row=row, column=1, value='OVERALL SCORE').font = Font(bold=True, size=12)
    overall = gap_analysis.get('overall_score', 0)
    cell = ws.cell(row=row, column=2, value=overall)
    cell.font = Font(bold=True, size=14)
    cell.fill = score_fill(overall)

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 60


def create_keyword_sheet(wb, keyword_analysis):
    """Sheet 2: Keyword Analysis — all keywords with priority tiers."""
    ws = wb.create_sheet("Keyword Analysis")

    headers = ['Keyword', 'Search Volume', 'Competing Products', 'Organic Rank',
               'Competitor Title Usage', 'Competitor Bullet Usage', 'Priority Score', 'Tier']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    tier_fills = {
        'title_priority': PatternFill(start_color="B4C6E7", fill_type="solid"),
        'bullet_priority': PatternFill(start_color="D9E2F3", fill_type="solid"),
        'backend_priority': PatternFill(start_color="E2EFDA", fill_type="solid"),
        'skip': PatternFill(start_color="F2F2F2", fill_type="solid"),
    }

    keywords = keyword_analysis.get('keywords', [])
    for i, kw in enumerate(keywords[:200], 2):  # Limit to top 200
        ws.cell(row=i, column=1, value=kw.get('keyword', ''))
        ws.cell(row=i, column=2, value=kw.get('search_volume', 0))
        ws.cell(row=i, column=3, value=kw.get('competing_products', 0))
        ws.cell(row=i, column=4, value=kw.get('organic_rank', 0) or 'N/A')
        ws.cell(row=i, column=5, value=kw.get('competitor_title_usage', 0))
        ws.cell(row=i, column=6, value=kw.get('competitor_bullet_usage', 0))
        ws.cell(row=i, column=7, value=kw.get('priority_score', 0))
        tier = kw.get('tier', 'skip')
        cell = ws.cell(row=i, column=8, value=tier.replace('_', ' ').title())
        cell.fill = tier_fills.get(tier, tier_fills['skip'])

    # Column widths
    widths = [35, 15, 18, 14, 20, 20, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_competitor_comparison_sheet(wb, competitor_listings, our_listing):
    """Sheet 3: Competitor Listing Comparison."""
    ws = wb.create_sheet("Competitor Comparison")

    headers = ['ASIN', 'Brand', 'Title', 'Title Length', 'Bullet Count',
               'Avg Bullet Length', 'Image Count', 'Has A+', 'Price', 'Rating']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    # Our listing first (highlighted)
    our_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
    our_bullets = our_listing.get('bullets', [])
    row_data = [
        our_listing.get('asin', 'OUR ASIN'),
        our_listing.get('brand', 'Zoviro'),
        our_listing.get('title', ''),
        len(our_listing.get('title', '')),
        len(our_bullets),
        round(sum(len(b) for b in our_bullets) / max(len(our_bullets), 1)),
        our_listing.get('imageCount', our_listing.get('image_count', 0)),
        'Yes' if our_listing.get('hasAPlus', our_listing.get('has_a_plus')) else 'No',
        our_listing.get('price', ''),
        our_listing.get('rating', '')
    ]
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = our_fill
        cell.alignment = WRAP_ALIGN

    # Competitor listings
    listings = competitor_listings if isinstance(competitor_listings, list) else competitor_listings.get('listings', [])
    for i, comp in enumerate(listings, 3):
        comp_bullets = comp.get('bullets', [])
        row_data = [
            comp.get('asin', ''),
            comp.get('brand', ''),
            comp.get('title', ''),
            len(comp.get('title', '')),
            len(comp_bullets),
            round(sum(len(b) for b in comp_bullets) / max(len(comp_bullets), 1)) if comp_bullets else 0,
            comp.get('imageCount', comp.get('image_count', 0)),
            'Yes' if comp.get('hasAPlus', comp.get('has_a_plus')) else 'No',
            comp.get('price', ''),
            comp.get('rating', '')
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = WRAP_ALIGN

    # Column widths
    widths = [14, 20, 70, 12, 12, 16, 12, 10, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_optimized_copy_sheet(wb, optimized_copy):
    """Sheet 4: Optimized Copy — title, bullets, backend keywords."""
    ws = wb.create_sheet("Optimized Copy")

    # Title Section
    ws.cell(row=1, column=1, value='TITLE OPTIMIZATION').font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    ws.cell(row=2, column=1, value='Current Title:').font = Font(bold=True)
    title_data = optimized_copy.get('title', {})
    current = title_data.get('current', {})
    ws.cell(row=2, column=2, value=current.get('title', ''))
    ws.cell(row=2, column=4, value=f"{current.get('char_count', 0)} chars")

    ws.cell(row=3, column=1, value='Missing Keywords:').font = Font(bold=True)
    ws.cell(row=3, column=2, value=', '.join(current.get('keywords_missing', [])))

    guidelines = title_data.get('guidelines', {})
    ws.cell(row=5, column=1, value='Must Include:').font = Font(bold=True)
    ws.cell(row=5, column=2, value=', '.join(guidelines.get('must_include_keywords', [])))

    ws.cell(row=6, column=1, value='Should Include:').font = Font(bold=True)
    ws.cell(row=6, column=2, value=', '.join(guidelines.get('should_include_keywords', [])))

    ws.cell(row=7, column=1, value='Format:').font = Font(bold=True)
    ws.cell(row=7, column=2, value=guidelines.get('format', ''))

    # Bullet Section
    row = 9
    ws.cell(row=row, column=1, value='BULLET POINT OPTIMIZATION').font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:D{row}')

    bullet_data = optimized_copy.get('bullet_points', {})
    bullet_guidelines = bullet_data.get('guidelines', {})
    themes = bullet_guidelines.get('bullet_themes', [])

    row += 1
    ws.cell(row=row, column=1, value='Position').font = Font(bold=True)
    ws.cell(row=row, column=2, value='Theme').font = Font(bold=True)
    ws.cell(row=row, column=3, value='Keywords to Include').font = Font(bold=True)
    ws.cell(row=row, column=4, value='Current Bullet').font = Font(bold=True)
    style_header_row(ws, row, 4)

    current_bullets = bullet_data.get('current_bullets', [])
    for i, theme in enumerate(themes):
        row += 1
        ws.cell(row=row, column=1, value=theme.get('position', i+1))
        ws.cell(row=row, column=2, value=theme.get('theme', ''))
        ws.cell(row=row, column=3, value=', '.join(theme.get('keywords', [])))
        ws.cell(row=row, column=4, value=current_bullets[i] if i < len(current_bullets) else '(missing)')
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = WRAP_ALIGN

    # Backend Keywords Section
    row += 2
    ws.cell(row=row, column=1, value='BACKEND SEARCH TERMS').font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:D{row}')

    backend = optimized_copy.get('backend_search_terms', {})
    row += 1
    ws.cell(row=row, column=1, value='Bytes Used:').font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"{backend.get('bytes_used', 0)}/250")

    row += 1
    ws.cell(row=row, column=1, value='Suggested Keywords:').font = Font(bold=True)
    selected = backend.get('selected_keywords', [])
    kw_text = ' '.join(kw.get('keyword', '') if isinstance(kw, dict) else str(kw) for kw in selected)
    ws.cell(row=row, column=2, value=kw_text)

    # A+ Section
    row += 2
    ws.cell(row=row, column=1, value='A+ CONTENT RECOMMENDATIONS').font = Font(bold=True, size=14)
    aplus = optimized_copy.get('aplus_content', {})
    row += 1
    ws.cell(row=row, column=1, value='Status:').font = Font(bold=True)
    ws.cell(row=row, column=2, value=aplus.get('status', ''))

    modules = aplus.get('suggested_modules', aplus.get('recommendations', []))
    if isinstance(modules, list):
        for mod in modules:
            row += 1
            if isinstance(mod, dict):
                ws.cell(row=row, column=1, value=mod.get('module', ''))
                ws.cell(row=row, column=2, value=mod.get('purpose', ''))
            else:
                ws.cell(row=row, column=1, value=str(mod))

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 60


def create_action_items_sheet(wb, gap_analysis):
    """Sheet 5: Action Items — prioritized recommendations."""
    ws = wb.create_sheet("Action Items")

    headers = ['Priority', 'Area', 'Action', 'Expected Impact']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    priority_fills = {
        'High': PatternFill(start_color="FFC7CE", fill_type="solid"),
        'Medium': PatternFill(start_color="FFEB9C", fill_type="solid"),
        'Low': PatternFill(start_color="C6EFCE", fill_type="solid"),
    }

    recs = gap_analysis.get('recommendations', [])
    for i, rec in enumerate(recs, 2):
        priority = rec.get('priority', 'Medium')
        ws.cell(row=i, column=1, value=priority).fill = priority_fills.get(priority, priority_fills['Medium'])
        ws.cell(row=i, column=2, value=rec.get('area', ''))
        ws.cell(row=i, column=3, value=rec.get('action', ''))
        ws.cell(row=i, column=4, value=rec.get('expected_impact', ''))

        for col in range(1, 5):
            ws.cell(row=i, column=col).alignment = WRAP_ALIGN

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 50


def generate_xlsx(gap_analysis_path, keyword_analysis_path, optimized_copy_path,
                  our_listing_path, competitor_listings_path, output_path):
    """Main XLSX generation."""

    with open(gap_analysis_path) as f:
        gap_analysis = json.load(f)
    with open(keyword_analysis_path) as f:
        keyword_analysis = json.load(f)
    with open(optimized_copy_path) as f:
        optimized_copy = json.load(f)
    with open(our_listing_path) as f:
        our_listing = json.load(f)
    with open(competitor_listings_path) as f:
        competitor_listings = json.load(f)

    wb = Workbook()

    create_scorecard_sheet(wb, gap_analysis)
    create_keyword_sheet(wb, keyword_analysis)
    create_competitor_comparison_sheet(wb, competitor_listings, our_listing)
    create_optimized_copy_sheet(wb, optimized_copy)
    create_action_items_sheet(wb, gap_analysis)

    wb.save(output_path)
    print(f"\nListing Optimization XLSX saved to: {output_path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate Listing Optimization XLSX report')
    parser.add_argument('--gap-analysis', required=True)
    parser.add_argument('--keyword-analysis', required=True)
    parser.add_argument('--optimized-copy', required=True)
    parser.add_argument('--our-listing', required=True)
    parser.add_argument('--competitor-listings', required=True)
    parser.add_argument('--output', required=True)
    args = parser.parse_args()

    generate_xlsx(args.gap_analysis, args.keyword_analysis, args.optimized_copy,
                  args.our_listing, args.competitor_listings, args.output)
